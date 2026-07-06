"""binder_status.py -- build the binder status / sign-off sheet (Excel).

Sign-off status = each WPM folder-contents row's `signOffs[]` (NOT documentStatus/formStatus, which are
lifecycle state). Deep link = the row's `documentUrl` (KCForms -> Knowledge Coach; everything else ->
engagement app). Bridge-native: the module captures auth, walks folders via chrome_api_call folder_get
(or wpm.folder_get on the linked-tab fallback), then calls rows_from_folders + build_status_workbook here.
"""

def _deep_link(t, u):
    if not u:
        return ""
    if t == "KCForms":
        return "https://knowledgecoach.cchaxcess.com/" + u.lstrip("/")
    return "https://engagement.cchaxcess.com" + (u if u.startswith("/") else "/" + u)

def _signoff(so):
    if not so:
        return "Not signed off"
    parts = []
    for s in so:
        if isinstance(s, dict):
            who = s.get("staffInitials") or s.get("initials") or s.get("staffName") or s.get("staffId") or ""
            role = s.get("role") or s.get("type") or s.get("signOffType") or ""
            dt = (s.get("signOffDate") or s.get("date") or "")[:10]
            parts.append(" ".join(x for x in [role, who, dt] if x) or "signed")
        else:
            parts.append(str(s))
    return "; ".join(parts)

def rows_from_folders(folder_blocks):
    """folder_blocks: [{"fIdx","fName","items":[raw WPM folder_get rows]}, ...] -> flat row dicts."""
    out = []
    for b in folder_blocks:
        for o in (b.get("items") or []):
            if o.get("type") == "Folder":
                continue
            out.append({
                "fIdx": b.get("fIdx", ""), "fName": b.get("fName", ""),
                "type": o.get("type", ""), "name": o.get("name", ""),
                "idx": (o.get("index") or o.get("documentIndex") or ""),
                "signOffs": (o.get("signOffs") or []),
                "documentUrl": (o.get("documentUrl") or ""),
                "formStatus": (o.get("formStatus") or ""),
                "statusCode": (o.get("statusCode") or ""),
                "lastModifiedOn": (o.get("lastModifiedOn") or ""),
            })
    return out

def build_status_workbook(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    rows = sorted(rows, key=lambda r: ((r.get("fIdx") or "zzzz"), (r.get("idx") or "")))
    wb = Workbook(); ws = wb.active; ws.title = "Binder Status Map"
    ws.append(["Folder","Index","Type","Workpaper / Form","Sign-off Status","Form/File Status","Last Modified","Open"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(vertical="center")
    link = Font(color="0563C1", underline="single")
    for r in rows:
        fld = ((r.get("fIdx","") + " " + r.get("fName","")).strip()) if r.get("fIdx") else r.get("fName","")
        ws.append([fld, r.get("idx",""), r.get("type",""), r.get("name",""), _signoff(r.get("signOffs") or []),
                   (r.get("formStatus") or r.get("statusCode") or ""), (r.get("lastModifiedOn") or "")[:10], "Open"])
        cell = ws.cell(row=ws.max_row, column=8); u = _deep_link(r.get("type",""), r.get("documentUrl",""))
        if u:
            cell.hyperlink = u; cell.font = link
    for i, w in enumerate([26,10,11,52,18,16,12,7], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = "A1:H%d" % ws.max_row
    wb.save(out_path); return out_path

# <!-- END -->
