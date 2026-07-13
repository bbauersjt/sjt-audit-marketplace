"""TB backup package — pull TB (CY + optional PY) + groups + subgroups + funds and
emit one workbook: sterile per-period import sheets, an Account Map across years,
Groups, Fund Structure, and a consolidated all-fields tab.

Execution model (architecture.md): the PULLS are JS built by the existing builders
(`scripts.groups`, `scripts.funds`) and executed in the browser tab; this module
PARSES the results and ASSEMBLES the workbook locally with openpyxl. No Python HTTP.

Sign convention (endpoints/fp_trialbalance.json + modules/import-tb-format.md):
the FP API serves balances CREDITS-POSITIVE; TB import/export convention is
DEBITS-POSITIVE. Import sheets and Account Map balances are sign-FLIPPED to
debits-positive; the All Fields tab keeps the RAW API values (labeled as such).

Subgroup key: `account.financialSubGroup`. `account.subGroup` is always null
— never read it.

Typical flow (module: references/modules/tb-backup-package.md):
    from scripts import groups, funds, tb_backup
    js = groups.get_trialbalance_grouped_all(cid, list_id, period_id, "ls:fp")
    rows = tb_backup.parse_tb(js_result)              # one call per period
    ...
    tb_backup.build_workbook(out_path, "Engagement 2025", tb_by_period={"CY-FINAL": rows_cy,
        "PY-FINAL": rows_py}, groups=parsed_groups, fund_types=..., funds=...,
        fund_sub_types=...)
"""
from __future__ import annotations

import json
from typing import Any, Optional

# --- parsing ----------------------------------------------------------------


def parse_tb(js_result_str: str) -> list[dict]:
    """Parse the result of groups.get_trialbalance_grouped_all().

    That builder resolves to JSON: {status, accountCount, count, balances:[...]}.
    Raises if the page-merge came back short (count != accountCount) — a partial
    TB silently poisons every downstream sheet, so it's a hard stop (SKILL.md
    failure discipline: report, don't improvise around it).
    """
    outer = json.loads(js_result_str)
    if isinstance(outer, dict) and "body" in outer and "balances" not in outer:
        # tolerate a build_xhr_call-shaped result (single page, small TB)
        body = outer.get("body")
        outer = json.loads(body) if isinstance(body, str) else (body or {})
    status = outer.get("status")
    if status is not None and status != 200:
        raise RuntimeError(f"TB pull failed: status {status}")
    rows = outer.get("balances") or []
    expected = outer.get("accountCount")
    got = outer.get("count", len(rows))
    if expected is not None and got != expected:
        raise RuntimeError(
            f"TB pull is PARTIAL: merged {got} rows but accountCount={expected}. "
            "Re-pull with get_trialbalance_grouped_all (paginated) before building.")
    return normalize_tb_rows(rows)


def _acct(row: dict) -> dict:
    return row.get("account") or {}


def _cur(row: dict) -> dict:
    periods = row.get("periods") or {}
    return periods.get("current") or {}


def _pick_balance(row: dict) -> Optional[float]:
    """Final balance if present, else adjusted, else unadjusted (RAW API sign)."""
    cur = _cur(row)
    for k in ("finalBalance", "adjustedBalance", "unadjustedBalance"):
        v = cur.get(k)
        if v is not None:
            return v
    # live FP variant: a flat 'final' balance on the account itself
    fv = _acct(row).get("final")
    return _to_float(fv) if fv is not None else None


def flip_sign(value: Optional[float]) -> Optional[float]:
    """API credits-positive → import/export debits-positive."""
    if value is None:
        return None
    return -value if value else 0.0


# --- live-FP normalization ------------------------------------------
#
# The grouped get_trialbalance_grouped_all() shape is the validated one
# (account.accountNumber, periods.current.finalBalance as a number, group/fund as
# strings). A raw/live FP pull can instead carry account.number, comma-formatted
# string balances, and group/fund as OBJECTS — which silently poison every
# downstream sheet (blank account numbers, text balances, "[object Object]" groups).
# normalize_tb_rows() coerces the live variant into the expected shape and is a
# NO-OP on already-normalized rows. parse_tb() runs it on every pull.

def _to_float(v: Any) -> Optional[float]:
    """Coerce a comma/paren/$-formatted balance string to float. Passes numbers
    through; returns None for blank/unparseable."""
    if v is None or isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "").replace("$", "").strip()
    if s.startswith("-"):
        neg = True
        s = s[1:]
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _as_label(v: Any) -> Any:
    """A group/fund/subgroup field may arrive as a dict (live FP) or a string.
    Return a stable string label (name → index → str); pass strings through."""
    if isinstance(v, dict):
        for k in ("name", "groupName", "fundName", "financialSubGroup",
                  "subGroupName", "index", "id"):
            if v.get(k):
                return str(v[k])
        return ""
    return v


def normalize_tb_rows(rows: list[dict]) -> list[dict]:
    """In-place: make live-FP-variant rows match the grouped TB shape. No-op on
    already-normalized rows.
      - account.number          -> account.accountNumber (keeps both)
      - group/fund/subGroup objects -> string labels
      - account.final / periods.current.*Balance comma-strings -> float
    """
    for row in rows or []:
        a = row.get("account")
        if isinstance(a, dict):
            if a.get("accountNumber") in (None, "") and a.get("number") is not None:
                a["accountNumber"] = a.get("number")
            for k in ("group", "fund", "financialSubGroup", "subGroup"):
                if isinstance(a.get(k), dict):
                    a[k] = _as_label(a[k])
            if isinstance(a.get("final"), str):
                a["final"] = _to_float(a["final"])
        cur = (row.get("periods") or {}).get("current")
        if isinstance(cur, dict):
            for k in ("finalBalance", "adjustedBalance", "unadjustedBalance"):
                if isinstance(cur.get(k), str):
                    cur[k] = _to_float(cur[k])
    return rows


def _flatten(row: dict) -> dict:
    """One flat dict per account row — every field, RAW API values."""
    out: dict[str, Any] = {}
    for k, v in _acct(row).items():
        out[f"account.{k}"] = v
    periods = row.get("periods") or {}
    for pname, pdata in periods.items():
        if isinstance(pdata, dict):
            for k, v in pdata.items():
                out[f"periods.{pname}.{k}"] = v
    for k, v in row.items():
        if k not in ("account", "periods"):
            out[k] = v
    return out


# --- workbook ----------------------------------------------------------------

_BAL_FMT = '#,##0.00_);(#,##0.00)'


def build_workbook(out_path: str,
                   engagement_name: str,
                   tb_by_period: dict[str, list[dict]],
                   groups: Optional[list[dict]] = None,
                   fund_types: Optional[list[dict]] = None,
                   funds: Optional[list[dict]] = None,
                   fund_sub_types: Optional[list[dict]] = None,
                   include_all_fields: bool = True) -> dict:
    """Assemble the backup workbook. Returns a summary dict for verification.

    tb_by_period: {"CY-FINAL": rows, "PY-FINAL": rows} — insertion order is kept;
    the FIRST entry is treated as current year for the Account Map balance columns.

    Sheets:
      Import {label}  — one per period; sterile import sheet (import-tb-format.md
                        constraints: header row 1, no freeze panes, single balance
                        column debits-positive, text account numbers).
      Account Map     — account → group / financialSubGroup / fund + one balance
                        column per period (debits-positive).
      Groups          — grouping inventory (if provided).
      Fund Structure  — fund types / funds / sub-types (if provided).
      All Fields      — consolidated flat dump, RAW API values (credits-positive),
                        first period's rows annotated with every period's balances.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    summary: dict[str, Any] = {"engagement": engagement_name, "sheets": {}}

    def _sheet(title: str):
        ws = wb.create_sheet(title=title[:31])
        assert ws.freeze_panes is None  # import-tb-format: freeze panes = hard reject
        return ws

    # -- per-period sterile import sheets --------------------------------------
    for label, rows in tb_by_period.items():
        ws = _sheet(f"Import {label}")
        ws.append(["Account Number", "Account Name", "Balance"])
        for c in ws[1]:
            c.font = bold
        total = 0.0
        for row in rows:
            a = _acct(row)
            bal = flip_sign(_pick_balance(row))
            cell_row = [str(a.get("accountNumber") or ""), a.get("accountName"), bal]
            ws.append(cell_row)
            r = ws.max_row
            ws.cell(row=r, column=1).number_format = "@"          # text — keep leading zeros
            ws.cell(row=r, column=3).number_format = _BAL_FMT     # accounting format
            total += bal or 0.0
        summary["sheets"][f"Import {label}"] = {"rows": len(rows), "net": round(total, 2)}

    # -- account map ------------------------------------------------------------
    period_labels = list(tb_by_period.keys())
    by_number: dict[str, dict] = {}
    for label in period_labels:
        for row in tb_by_period[label]:
            a = _acct(row)
            num = str(a.get("accountNumber") or "")
            rec = by_number.setdefault(num, {
                "name": a.get("accountName"),
                "group": a.get("group"),
                "subgroup": a.get("financialSubGroup"),   # NOT a.get("subGroup") — always null
                "fund": a.get("fund"),
                "balances": {},
            })
            rec["balances"][label] = flip_sign(_pick_balance(row))
            # prefer current-year (first period) metadata when both exist
            if label == period_labels[0]:
                rec.update({"name": a.get("accountName"), "group": a.get("group"),
                            "subgroup": a.get("financialSubGroup"), "fund": a.get("fund")})

    ws = _sheet("Account Map")
    header = (["Account Number", "Account Name", "Group", "Subgroup", "Fund"]
              + [f"Balance {p}" for p in period_labels])
    ws.append(header)
    for c in ws[1]:
        c.font = bold
    for num in sorted(by_number):
        rec = by_number[num]
        ws.append([num, rec["name"], rec["group"], rec["subgroup"], rec["fund"]]
                  + [rec["balances"].get(p) for p in period_labels])
        r = ws.max_row
        ws.cell(row=r, column=1).number_format = "@"
        for i in range(len(period_labels)):
            ws.cell(row=r, column=6 + i).number_format = _BAL_FMT
    summary["sheets"]["Account Map"] = {"rows": len(by_number)}

    # -- groups -------------------------------------------------------------------
    if groups:
        ws = _sheet("Groups")
        keys = sorted({k for g in groups for k in g.keys()})
        # put the human columns first when present
        front = [k for k in ("index", "name", "groupName", "financialGroupId") if k in keys]
        keys = front + [k for k in keys if k not in front]
        ws.append(keys)
        for c in ws[1]:
            c.font = bold
        for g in groups:
            ws.append([_cell_safe(g.get(k)) for k in keys])
        summary["sheets"]["Groups"] = {"rows": len(groups)}

    # -- fund structure -------------------------------------------------------------
    if fund_types or funds or fund_sub_types:
        ws = _sheet("Fund Structure")
        r = 1
        for title, data in (("FUND TYPES", fund_types), ("FUNDS", funds),
                            ("FUND SUB-TYPES", fund_sub_types)):
            if not data:
                continue
            ws.cell(row=r, column=1, value=title).font = bold
            r += 1
            keys = sorted({k for d in data for k in d.keys()})
            for ci, k in enumerate(keys, start=1):
                ws.cell(row=r, column=ci, value=k).font = bold
            r += 1
            for d in data:
                for ci, k in enumerate(keys, start=1):
                    ws.cell(row=r, column=ci, value=_cell_safe(d.get(k)))
                r += 1
            r += 1  # blank spacer row between sections
        summary["sheets"]["Fund Structure"] = {
            "fund_types": len(fund_types or []), "funds": len(funds or []),
            "fund_sub_types": len(fund_sub_types or [])}

    # -- consolidated all-fields tab (default ON) ------------------------------------
    if include_all_fields:
        ws = _sheet("All Fields")
        flat_rows = []
        for label, rows in tb_by_period.items():
            for row in rows:
                fr = _flatten(row)
                fr["_period"] = label
                flat_rows.append(fr)
        keys = sorted({k for fr in flat_rows for k in fr.keys()})
        front = [k for k in ("_period", "account.accountNumber", "account.accountName",
                             "account.group", "account.financialSubGroup",
                             "account.fund") if k in keys]
        keys = front + [k for k in keys if k not in front]
        ws.append(["RAW API VALUES — credits-positive (fp_trialbalance.json); "
                   "import sheets are the sign-flipped truth"])
        ws.cell(row=1, column=1).font = bold
        ws.append(keys)
        for c in ws[2]:
            c.font = bold
        for fr in flat_rows:
            ws.append([_cell_safe(fr.get(k)) for k in keys])
        summary["sheets"]["All Fields"] = {"rows": len(flat_rows)}

    wb.save(out_path)
    summary["path"] = out_path
    return summary


def _cell_safe(v):
    """openpyxl rejects dict/list cell values — JSON-encode the odd nested field."""
    if isinstance(v, (dict, list)):
        return json.dumps(v)
    return v


def verify_workbook(out_path: str, tb_by_period: dict[str, list[dict]]) -> dict:
    """Re-open the produced file and re-check the import-format constraints +
    row counts. Returns {ok, problems:[...]} — report problems, don't patch silently."""
    from openpyxl import load_workbook
    problems = []
    wb = load_workbook(out_path)
    for label, rows in tb_by_period.items():
        name = f"Import {label}"[:31]
        if name not in wb.sheetnames:
            problems.append(f"missing sheet {name}")
            continue
        ws = wb[name]
        if ws.freeze_panes not in (None, "A1"):
            problems.append(f"{name}: freeze panes set ({ws.freeze_panes}) — CCH hard-rejects")
        if [c.value for c in ws[1]][:3] != ["Account Number", "Account Name", "Balance"]:
            problems.append(f"{name}: header row is not row 1 / wrong headers")
        if ws.max_row - 1 != len(rows):
            problems.append(f"{name}: {ws.max_row - 1} data rows != {len(rows)} pulled rows")
    return {"ok": not problems, "problems": problems}
# <!-- END -->
