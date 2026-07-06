"""Cross-reference extraction & resolution for KC forms.

Pure Python — operates on already-fetched form JSON. No HTTP.
"""
import json
import re
from pathlib import Path

import openpyxl

SKILL_ROOT = Path(__file__).parent.parent
DENYLIST_PATH = SKILL_ROOT / "references" / "config" / "xref_denylist.json"
XREFS_DIR = SKILL_ROOT / "references" / "data" / "engagement-xrefs"

FID_RE = re.compile(r"\b([A-Z]{3}-\d{3,4}[A-Z]?)\b")


def _load_denylist() -> list[re.Pattern]:
    data = json.loads(DENYLIST_PATH.read_text())
    return [re.compile(p) for p in data.get("patterns", [])]


def extract_form_refs(form_json: dict, exclude: str | None = None) -> set[str]:
    """Find every KC form reference in a form's JSON, minus denylist matches.

    exclude: the current form's ID (so it doesn't reference itself).
    """
    blob = json.dumps(form_json)
    raw = set(FID_RE.findall(blob))
    if exclude:
        raw.discard(exclude)
    deny = _load_denylist()
    return {fid for fid in raw if not any(p.match(fid) for p in deny)}


def load_engagement_form_index(engagement_slug: str) -> dict[str, str]:
    """Read references/data/engagement-xrefs/{slug}.xlsx — 'Form Index' sheet.

    Returns {form_id: assigned_index}. Empty dict if file missing.
    """
    path = XREFS_DIR / f"{engagement_slug}.xlsx"
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Form Index" not in wb.sheetnames:
        return {}
    ws = wb["Form Index"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        fid = str(row[0]).strip()
        idx = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if fid:
            out[fid] = idx
    return out


def resolve(refs: set[str], form_index: dict[str, str]) -> tuple[list[dict], list[str]]:
    """Split a set of refs into (resolved, unresolved).

    resolved: [{fid, idx}, ...] in sorted order.
    unresolved: [fid, ...] in sorted order — these need a user decision.
    """
    resolved, unresolved = [], []
    for fid in sorted(refs):
        if fid in form_index and form_index[fid]:
            resolved.append({"fid": fid, "idx": form_index[fid]})
        else:
            unresolved.append(fid)
    return resolved, unresolved


def enrich_text(text: str, form_index: dict[str, str]) -> str:
    """Inline-annotate KC form references in free text with their assigned indexes.

    'see KBA-302 for ...' -> 'see KBA-302 (at 0102) for ...'
    Leaves unresolved references untouched.
    """
    def _sub(m):
        fid = m.group(1)
        idx = form_index.get(fid)
        return f"{fid} (at {idx})" if idx else fid
    return FID_RE.sub(_sub, text)
# <!-- END -->
