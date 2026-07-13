"""Read kc-forms-catalog-rich.xlsx and assemble Add-Forms POST bodies.

Pure Python, runs in the sandbox. Does not hit any HTTP endpoint.

ACTUAL xlsx columns (sheet "Catalog"):
    Form ID | Form Name | Category | Title | Reference Tag | Data Binding Key
    | Description | Tags | Title GUID | Add-Form Constants
"Add-Form Constants" packs the POST constants as "rf=3;copy=15;upd=15;mpk=KEY".
"""
import openpyxl
from pathlib import Path

SKILL_ROOT = Path(__file__).parent.parent
CATALOG_PATH = SKILL_ROOT / "references" / "data" / "kc-forms-catalog-rich.xlsx"

# xlsx column -> Add-Forms POST body field
COLUMN_TO_BODY_FIELD = {
    "Form Name": "name",
    "Description": "description",
    "Category": "group",
    "Reference Tag": "referenceTag",
    "Data Binding Key": "dataBindingKey",
    "Title GUID": "titleID",
}


def load_catalog() -> list[dict]:
    """Return the full catalog as a list of dicts keyed by the REAL column headers."""
    wb = openpyxl.load_workbook(CATALOG_PATH, data_only=True, read_only=True)
    ws = wb["Catalog"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
        rows.append(d)
    return rows


def lookup_by_form_id(form_id: str, title_guid: str | None = None) -> dict | None:
    """Find one form row by Form ID (e.g. "AID-201"), optionally pinned to a title.

    This is the PRIMARY lookup — Form ID is what plans and users speak in.
    """
    for row in load_catalog():
        if row.get("Form ID") == form_id and (title_guid is None or row.get("Title GUID") == title_guid):
            return row
    return None


def lookup_by_reference_tag(reference_tag: str, title_guid: str | None = None) -> dict | None:
    """Find one form row by Reference Tag (and optionally Title GUID)."""
    for row in load_catalog():
        if row.get("Reference Tag") == reference_tag and (title_guid is None or row.get("Title GUID") == title_guid):
            return row
    return None


def resolve_title_guid(title_guid: str) -> str | None:
    """Human-readable KC title for a Title GUID — THE authoritative entity-type signal.

    GetBinder's `lastUsedTitleGuid` (set when the user picks the KC title at binder
    creation; never null on a real engagement) resolves here to e.g.
    '2025 - Knowledge-Based Audits of Governmental Entities'. Use THIS for entity
    type — never the client name, and never KC form ref prefixes (KBA-1xx refs are
    shared across entity types). Corroborate with TB account structure
    (see setup-binder-from-index.md Phase 0).
    """
    g = (title_guid or "").strip().lower()
    if not g:
        return None
    for row in load_catalog():
        if row.get("Title GUID", "").strip().lower() == g:
            return row.get("Title") or None
    return None


def _parse_add_form_constants(packed: str) -> dict:
    """'rf=3;copy=15;upd=15;mpk=PRIMARYMAJORCONTAINERLIST' -> POST body fields."""
    out = {"rfSettings": "", "copySettings": "", "updateSettings": "", "majorProgramKey": ""}
    key_map = {"rf": "rfSettings", "copy": "copySettings", "upd": "updateSettings", "mpk": "majorProgramKey"}
    for part in (packed or "").split(";"):
        if "=" not in part:
            continue
        k, _, v = part.partition("=")
        field = key_map.get(k.strip())
        if not field:
            continue
        v = v.strip()
        out[field] = int(v) if v.isdigit() else v
    return out


def build_add_forms_body(rows: list[dict]) -> list[dict]:
    """Assemble the POST body for kc_add_forms from a list of catalog rows.

    Maps the REAL xlsx columns to body fields and unpacks "Add-Form Constants".
    Other fields are constants or placeholder GUIDs the server replaces —
    documented in references/endpoints/kc_add_forms.json.
    """
    PLACEHOLDER_GUID = "00000000-0000-0000-0000-000000000000"
    bodies = []
    for r in rows:
        body = {field: r.get(col, "") for col, field in COLUMN_TO_BODY_FIELD.items()}
        body.update(_parse_add_form_constants(r.get("Add-Form Constants", "")))
        body.update({
            "workpaperId": PLACEHOLDER_GUID,
            "binderTitleId": r.get("Title GUID", ""),
            "isDeleted": False,
            "isVisible": True,
            "isLocked": False,
            "index": "",  # ignored by server — set via wpm_set_index after move
        })
        bodies.append(body)
    return bodies


# --- Single Audit (S-suffix) title support ------------------------------
SA_TITLE_GUID = "531eb5ad-5eae-4f12-ac51-ea998bb8472e"   # Knowledge-Based Single Audits SAS.2024.1
SA_TITLE_NAME = "Knowledge-Based Single Audits"


def load_sa_title_forms() -> list[dict]:
    """Static fallback list of the Single Audits title's 84 forms
    (references/data/sa-title-forms.tsv).

    S-suffix forms are NOT in the binder's own title — they live in the separate
    Single Audits title and are added CROSS-TITLE (same add_forms POST, SA
    titleGuid). Prefer a LIVE pull when a KC tab is open:
    GET /api/binder/GetWorkpaperListForAddForms/{binderGuid}/{SA_TITLE_GUID}
    (ls:kc auth — see endpoints/kc_title_library.json); this TSV is the offline
    fallback. Returns [{fid, reference_tag, name, section}].
    """
    import pathlib
    tsv = pathlib.Path(__file__).resolve().parent.parent / "references" / "data" / "sa-title-forms.tsv"
    rows = []
    for line in tsv.read_text().splitlines():
        parts = line.split("\t")
        if len(parts) >= 4:
            rows.append({"fid": parts[0], "reference_tag": parts[1],
                         "name": parts[2], "section": parts[3]})
    return rows

# <!-- END -->
