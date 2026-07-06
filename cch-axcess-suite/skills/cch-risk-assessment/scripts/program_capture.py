#!/usr/bin/env python3
"""Re-runnable capture of AUD-8xx program TAILORING QUESTIONS into program_question.

WHY THIS EXISTS: the CCH tailoring-question set changes every year (new title
release). This module makes the yearly re-capture reproducible and auditable
instead of a pile of ad-hoc snippets. The seed it produces is data/seed/
program_question.csv, queried at runtime by program.py (NOT by reading MDs).

SPLIT: the auth'd HTTP half runs in the user's Chrome tab (JS below); the parse/
build/write half runs here in the sandbox. They hand off via a JSON file the
browser downloads (the javascript_tool return cap is too small/erratic to stream
~200 rows -- the Blob-download bridge is the reliable transfer).

=========================== CAPTURE PROTOCOL ===========================
Run once per audit title per year. NPO title GUID = dcf6ee8e-6256-48d6-a3bc-91939f4d7792.

1. SANDBOX: body = build_add_bodies(CATALOG_PATH, TITLE_GUID, at="NPO")
2. CHROME (a KC tab on the engagement, tokens in localStorage):
     run JS_ADD(body_json)        -> POSTs /api/binder/{eng}; adds all 22 programs
     run JS_CAPTURE(eng)          -> GetBinder, GET each AUD-8xx, parse .{AREA}.TailoringQuestions,
                                     download 'cch_tailoring_questions.json' (flat [[AUD,key,text],...])
3. Mount the user's Downloads folder; then SANDBOX:
     write_questions("<Downloads>/cch_tailoring_questions.json", at="NPO")
4. SANDBOX: python build_db.py   (rebuild) ; python program.py --list  (validate counts)

VALIDATION: program.py --list should show ~20 programs (JE2 + ESTIMATES carry 0
tailoring questions by design). Spot-check a program against the live CCH form.

NOTE: excludes .{AREA}.ConTestingEffControlQuestion (controls-testing Y/N) -- empty
Question text + generic shared keys, not area-specific drivers. Add a second qtype
pass here if the firm decides to track them.
========================================================================
"""
import csv, json, os
from _db import SEED_DIR

CATALOG_PATH = ("/sessions/brave-intelligent-franklin/mnt/.claude/skills/"
                "cch-axcess/references/data/kc-forms-catalog-rich.xlsx")  # cch-axcess sibling skill; repoint per install
NPO_TITLE = "dcf6ee8e-6256-48d6-a3bc-91939f4d7792"
# audit_type -> CCH title GUID (from kc-forms-catalog-rich.xlsx). CNS/HOA deferred.
TITLES = {
    "NPO": "dcf6ee8e-6256-48d6-a3bc-91939f4d7792",  # 2026 Not-for-Profit
    "ASB": "c0b36674-ffc7-4e0c-be8d-d24deab3c509",  # 2025 Commercial Entities (HOA rides on this)
    "EBP": "1b32bc1a-e119-4cc9-91fe-b2cb22f4e966",  # 2025 Employee Benefit Plans
    "ALG": "3ac49bd1-97fa-49cf-b332-f95f529c7b59",  # 2025 Governmental Entities
    "CNS": "1ea4a73d-9c93-4f0c-ae63-d4940ec0c6fa",  # 2025 Construction Contractors (future)
}
_CONST = {"rf":3,"copy":15,"upd":15,"mpk":"PRIMARYMAJORCONTAINERLIST"}


def aud_to_area(at="NPO"):
    """Map {aud_form: binding_key} for a title, from areas.csv (source of truth)."""
    out={}
    with open(os.path.join(SEED_DIR,"areas.csv"),newline="") as f:
        for r in csv.DictReader(f):
            if r["audit_type"]==at and r["aud_form"]:
                out[r["aud_form"]]=r["binding_key"]
    return out


def build_add_bodies(catalog_path, title_id, at="NPO"):
    """Read kc-forms-catalog-rich.xlsx and assemble the Add-Forms POST body for
    every AUD-8xx area program of this title. Returns a list of dicts."""
    import openpyxl
    auds=set(aud_to_area(at)) | {f"AUD-{n}" for n in range(801,823)}
    wb=openpyxl.load_workbook(catalog_path,data_only=True,read_only=True)
    ws=wb["Catalog"]
    hdr=[str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1,max_row=1))]
    idx={h:i for i,h in enumerate(hdr)}
    PH="00000000-0000-0000-0000-000000000000"
    out=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not any(row): continue
        g=lambda c:(str(row[idx[c]]).strip() if c in idx and row[idx[c]] is not None else "")
        if g("Form ID") in auds and g("Title GUID")==title_id:
            d=dict(_CONST)
            for kv in (g("Add-Form Constants") or "").split(";"):
                if "=" in kv: k,v=kv.split("=",1); d[k.strip()]=v.strip()
            out.append({"name":g("Form Name"),"description":g("Description"),
                "group":g("Category"),"referenceTag":g("Reference Tag"),
                "dataBindingKey":g("Data Binding Key"),"titleID":title_id,
                "majorProgramKey":d["mpk"],"rfSettings":int(d["rf"]),
                "copySettings":int(d["copy"]),"updateSettings":int(d["upd"]),
                "workpaperId":PH,"binderTitleId":title_id,"isDeleted":False,
                "isVisible":True,"isLocked":False,"index":""})
    return out


def write_questions(flat_json_path, at="NPO"):
    """Write program_question.csv rows from the captured flat JSON [[AUD,q_id,text],...].
    Preserves rows of other audit_types; rewrites only this `at`."""
    a2a=aud_to_area(at)
    data=json.load(open(flat_json_path))
    csv_path=os.path.join(SEED_DIR,"program_question.csv")
    hdr=["audit_type","area","program","q_id","seq","qtype","question_text",
         "default_answer","autonomy","look_for","notes"]
    keep=[]
    if os.path.exists(csv_path):
        with open(csv_path,newline="") as f:
            r=csv.reader(f); next(r,None)
            keep=[row for row in r if row and row[0]!=at]
    seqs={}; new=[]
    for aud,q_id,text in data:
        area=a2a.get(aud,aud); seqs[area]=seqs.get(area,0)+1
        new.append([at,area,aud,q_id,seqs[area],"tailoring",text,"","","",""])
    with open(csv_path,"w",newline="") as f:
        w=csv.writer(f); w.writerow(hdr); w.writerows(keep); w.writerows(new)
    return len(new)


# ---- Chrome-side JS (run via mcp__Claude_in_Chrome__javascript_tool on a KC tab) ----

def JS_ADD(body_json_str):
    """JS: POST the add-forms body to /api/binder/{eng}. Pass json.dumps(build_add_bodies(...))."""
    return ("(async()=>{const at=localStorage.getItem('kc.accessToken'),it=localStorage.getItem('kc.idToken');"
            "if(!at)return'NO_TOKEN';const H={Authorization:'Bearer '+at,IdToken:it,Accept:'application/json','Content-Type':'application/json'};"
            "const eng=location.pathname.split('/binder/')[1].split('/')[0];"
            "const r=await fetch(`https://knowledgecoach.cchaxcess.com/api/binder/${eng}`,{method:'POST',headers:H,credentials:'include',body:JSON.stringify("+body_json_str+")});"
            "const j=await r.json();const a=j.result||j;return JSON.stringify({status:r.status,n:Array.isArray(a)?a.length:0});})()")

JS_CAPTURE = (r"""
(async()=>{
  const at=localStorage.getItem('kc.accessToken'),it=localStorage.getItem('kc.idToken');
  if(!at)return'NO_TOKEN';
  const H={Authorization:'Bearer '+at,IdToken:it,Accept:'application/json','Content-Type':'application/json'};
  const eng=location.pathname.split('/binder/')[1].split('/')[0];
  const P=v=>typeof v==='string'?JSON.parse(v||'[]'):(v||[]);
  const rp=o=>Array.isArray(o.renderProperties)?o.renderProperties:Object.values(o.renderProperties||{});
  const qtext=ps=>{const x=ps.find(p=>(p.key||'').toLowerCase()==='question');return x?(x.value||''):'';};
  let b=await(await fetch(`https://knowledgecoach.cchaxcess.com/api/binder/GetBinder/${eng}`,{headers:H,credentials:'include'})).json();
  let wps=(b.result&&(b.result.workpapers||b.result.Workpapers))||[];
  let prog=wps.filter(w=>/AUD-8\d\d/i.test(w.name||w.workpaperName||'')).map(w=>{let m=(w.name||w.workpaperName||'').match(/AUD-(8\d\d)/i);return['AUD-'+m[1],w.workpaperId||w.id];});
  const flat=[];
  for(const[aud,wp]of prog){
    let f=await(await fetch(`https://knowledgecoach.cchaxcess.com/api/Workpaper/${eng}/${wp}`,{headers:H,credentials:'include'})).json();
    let c=P(f.result.collections).find(x=>/\.TailoringQuestions$/i.test(x.path||x.key||''));
    if(c)for(const o of(c.objectList||[]))flat.push([aud,o.key,qtext(rp(o)).replace(/\s+/g,' ').trim()]);
  }
  const blob=new Blob([JSON.stringify(flat)],{type:'application/json'});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='cch_tailoring_questions.json';
  document.body.appendChild(a);a.click();a.remove();
  return 'downloaded cch_tailoring_questions.json: '+flat.length+' questions, '+prog.length+' programs';
})()
""")
