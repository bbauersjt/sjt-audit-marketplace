"""Extract the trial balance, groupings, fund structure, and AJEs from a
CaseWare Working Papers folder copy - no CaseWare needed.

Reads the engagement's FoxPro DBF tables:
  <Client>am.dbf - account master: AC_NO, AC_DESC, SCHEDULE (leadsheet/LS
                   grouping), GROUP2..GROUP9 (extra groupings - GROUP2 is the
                   firm 4-digit index where used), ENTITY (fund key), TYPE
                   (B=balance sheet, I=income stmt), SIGN
  <Client>CE.dbf - consolidation entities = fund structure: ABBR (fund no.),
                   NAME, TYPE (M=master, C=child), ENTITY sort key. The ENTITY
                   key is hierarchical - a child's key starts with its parent's
                   key - so the consolidation tree (which children roll up to
                   which master) is recovered by prefix matching. Masters below
                   the root map to CCH fund TYPES; children map to CCH funds.
  <Client>bl.dbf - balances: ID 'O' + BUCKET 'O' = UNADJUSTED year-end balance;
                   ID 'X' + BUCKET 'Y' = journal-entry totals by TYPE
                   (N=booked normal AJEs, L=reclass, U=unrecorded/passed);
                   YEAR '0' = current file year, '1'..'4' = years back.
                   Verified against a signed FS: final = unadjusted + type N.
  <Client>gl.dbf - current-year journal entry detail lines
  <Client>MP.dbf - grouping/mapping definitions: ID '1' rows carry the L/S
                   (SCHEDULE) legend - MAP_NO is the grouping code, AC_DESC its
                   name. Verified 100% L/S coverage across six firm files; the
                   GROUP2..GROUP9 structures (ID '2'+) exist but are unnamed.

Usage:
  python cw_tb_extract.py "<engagement folder>" [--out DIR]

Writes _tb_extract.xlsx (sheets: TB, Funds, AJEs) and _tb_extract.json into
the top of the engagement folder (or --out). Amounts keep CaseWare's natural
sign (credits negative). CCH import formatting is trial-balance-prep's job -
this script only gets the data out clean.
"""
import argparse
import datetime
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cw_crosswalk import read_dbf, find_sh_table  # noqa: E402


def find_table(folder, suffix):
    import glob
    hits = [p for p in glob.glob(os.path.join(folder, f'*{suffix}.dbf'))
            + glob.glob(os.path.join(folder, f'*{suffix}.DBF'))
            if os.sep + '_Sync' + os.sep not in p]
    return hits[0] if hits else None


def load(folder, suffix):
    path = find_table(folder, suffix)
    if not path:
        return []
    fpt = None
    for ext in ('.fpt', '.FPT'):
        cand = os.path.splitext(path)[0] + ext
        if os.path.exists(cand):
            fpt = cand
            break
    return [r for r in read_dbf(path, fpt) if not r['_deleted']]


def num(s):
    try:
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def extract(folder):
    accounts = load(folder, 'am')
    entities = load(folder, 'CE')
    balances = load(folder, 'bl')
    jelines = load(folder, 'gl')
    mappings = load(folder, 'MP')

    # grouping legends: MP ID '1' = the L/S (SCHEDULE) structure; higher IDs
    # are the GROUP2+ structures (present but unnamed in firm files)
    legends = {}
    for m in mappings:
        code, name = m['MAP_NO'].strip(), m['AC_DESC'].strip()
        if code:
            legends.setdefault(m['ID'], {})[code] = name
    ls_names = legends.get('1', {})

    # entity keys are hierarchical: child key = parent key + suffix
    keys = [e['ENTITY'] for e in entities]
    by_key = {e['ENTITY']: e for e in entities}

    def parent_key(k):
        pref = [o for o in keys if o != k and k.startswith(o)]
        return max(pref, key=len) if pref else None

    parents = {k: parent_key(k) for k in keys}
    roots = {k for k, p in parents.items() if p is None}

    def fund_type_key(k):
        p = parents.get(k)
        while p is not None and p not in roots:
            if by_key[p]['TYPE'] == 'M':
                return p
            p = parents.get(p)
        return None

    funds = {}
    for e in entities:
        k = e['ENTITY']
        ftk = fund_type_key(k)
        pk = parents.get(k)
        funds[k] = {
            'fund': e['ABBR'], 'name': e['NAME'],
            'type': ('root' if k in roots else
                     {'M': 'master', 'C': 'child'}.get(e['TYPE'], e['TYPE'])),
            'parent': by_key[pk]['ABBR'] if pk else '',
            'fund_type': by_key[ftk]['ABBR'] if ftk else '',
            'fund_type_name': by_key[ftk]['NAME'] if ftk else '',
            'year_end': e.get('PER', ''),
        }

    # balances keyed by (account, entity): unadjusted by year, JE totals by (year, type)
    unadj, jetot = {}, {}
    for b in balances:
        key = (b['NUMBER'], b['ENTITY'])
        yr = b['YEAR']
        if b['ID'] == 'O' and b['BUCKET'] == 'O':
            unadj.setdefault(key, {})[yr] = num(b['AMT'])
        elif b['ID'] == 'X' and b['BUCKET'] == 'Y':
            jetot.setdefault(key, {})[(yr, b['TYPE'])] = num(b['AMT'])

    rows = []
    for a in accounts:
        if a['AC_NO'] == 'NETINC':
            continue  # CaseWare's computed net-income row, not a real account
        key = (a['AC_NO'], a['ENTITY'])
        u = unadj.get(key, {})
        j = jetot.get(key, {})
        fund = funds.get(a['ENTITY'], {})
        row = {
            'fund': fund.get('fund', ''),
            'fund_name': fund.get('name', ''),
            'account': a['AC_NO'],
            'description': a['AC_DESC'],
            'leadsheet': a['SCHEDULE'],
            'leadsheet_name': ls_names.get(a['SCHEDULE'].strip(), ''),
            'group2': a['GROUP2'],
            'group3': a['GROUP3'],
            'map_no': a.get('PAC_NO', ''),
            'type': {'B': 'BS', 'I': 'IS'}.get(a['TYPE'], a['TYPE']),
            'sign': a['SIGN'],
            'cy_unadjusted': u.get('0', 0.0),
            'cy_aje': j.get(('0', 'N'), 0.0),
            'cy_reclass': j.get(('0', 'L'), 0.0),
            'cy_passed': j.get(('0', 'U'), 0.0),
            'cy_final': u.get('0', 0.0) + j.get(('0', 'N'), 0.0),
        }
        for n in ('1', '2', '3', '4'):
            row[f'py{n}_final'] = u.get(n, 0.0) + j.get((n, 'N'), 0.0)
        rows.append(row)

    ajes = []
    for je in jelines:
        if set(je.get('AC_NO', '')) == {'Z'}:
            continue  # JE header/control line, not a posting
        ajes.append({
            'date': je.get('DATE', ''), 'ref': je.get('REFNO', ''),
            'type': je.get('DTYPE', ''), 'description': je.get('JE_DESC', ''),
            'account': je.get('AC_NO', ''), 'entity_fund': funds.get(je.get('ENTITY', ''), {}).get('fund', ''),
            'amount': num(je.get('NET', '')),
        })
    used = {a['SCHEDULE'].strip() for a in accounts if a['SCHEDULE'].strip()}
    groupings = [{'code': c, 'name': n, 'used': c in used}
                 for c, n in sorted(ls_names.items()) if n]
    funds_out = [funds[k] for k in sorted(funds)]  # ENTITY key order = tree order
    return rows, funds_out, ajes, groupings


TB_COLS = ['fund', 'fund_name', 'account', 'description', 'leadsheet',
           'leadsheet_name', 'group2', 'group3', 'map_no', 'type', 'sign',
           'cy_unadjusted', 'cy_aje', 'cy_reclass', 'cy_passed', 'cy_final',
           'py1_final', 'py2_final', 'py3_final', 'py4_final']
TB_HEADERS = ['Fund', 'Fund Name', 'Account', 'Description', 'L/S', 'L/S Name',
              'Group 2', 'Group 3', 'Map No', 'BS/IS', 'Sign', 'CY Unadjusted',
              'CY AJEs', 'CY Reclass', 'CY Passed', 'CY Final',
              'PY1 Final', 'PY2 Final', 'PY3 Final', 'PY4 Final']


def write_xlsx(path, rows, funds, ajes, groupings):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()

    def style_header(ws):
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='44546A')
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    ws = wb.active
    ws.title = 'TB'
    ws.append(TB_HEADERS)
    for r in rows:
        ws.append([r[k] for k in TB_COLS])
    for col in range(12, 21):
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].number_format = '#,##0.00'
    for i, w in enumerate([8, 32, 14, 42, 8, 30, 10, 10, 10, 7, 6] + [14] * 9, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws)

    wsg = wb.create_sheet('Groupings')
    wsg.append(['L/S Code', 'Name', 'Used by an account'])
    for g in groupings:
        wsg.append([g['code'], g['name'], 'Y' if g['used'] else ''])
    for i, w in enumerate([10, 45, 18], 1):
        wsg.column_dimensions[get_column_letter(i)].width = w
    if groupings:
        style_header(wsg)

    ws2 = wb.create_sheet('Funds')
    ws2.append(['Fund', 'Name', 'Type', 'Parent', 'Fund Type', 'Fund Type Name', 'Year End'])
    for f in funds:
        ws2.append([f['fund'], f['name'], f['type'], f['parent'],
                    f['fund_type'], f['fund_type_name'], f['year_end']])
    for i, w in enumerate([10, 45, 10, 10, 10, 35, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    if funds:
        style_header(ws2)

    ws3 = wb.create_sheet('AJEs')
    ws3.append(['Date', 'Ref', 'Type', 'Description', 'Account', 'Fund', 'Amount'])
    for a in ajes:
        ws3.append([a['date'], a['ref'], a['type'], a['description'],
                    a['account'], a['entity_fund'], a['amount']])
    for cell in ws3.iter_rows(min_row=2, min_col=7, max_col=7):
        cell[0].number_format = '#,##0.00'
    for i, w in enumerate([11, 10, 6, 40, 14, 8, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    if ajes:
        style_header(ws3)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('folder')
    ap.add_argument('--out', default=None)
    args = ap.parse_args()
    folder = os.path.abspath(args.folder)
    if not find_table(folder, 'am'):
        print("SKIP: no *am.dbf found - need a full CaseWare folder copy")
        sys.exit(1)
    rows, funds, ajes, groupings = extract(folder)
    name = os.path.basename(folder.rstrip('\\/')).replace(' (Sync)', '')
    out = args.out or folder
    os.makedirs(out, exist_ok=True)
    meta = {
        'engagement': name,
        'generated': datetime.datetime.now().isoformat(timespec='seconds'),
        'note': ('Trial balance extracted from the CaseWare DBF tables. Amounts keep '
                 "CaseWare's natural sign (credits negative). cy_final = unadjusted + "
                 'booked normal AJEs (verified against signed FS). YEAR columns: cy = '
                 'the file year, py1..py4 = 1-4 years back. Reclass (L) and passed (U) '
                 'entries reported separately, not in cy_final. groupings is the '
                 "full L/S legend from MP.dbf (code -> name); leadsheet_name on "
                 "each TB row resolves that account's L/S code."),
        'groupings': groupings,
        'funds': funds,
        'trial_balance': rows,
        'ajes': ajes,
    }
    json_path = os.path.join(out, '_tb_extract.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(meta, f, indent=1)
    xlsx_path = os.path.join(out, '_tb_extract.xlsx')
    write_xlsx(xlsx_path, rows, funds, ajes, groupings)
    cy = sum(r['cy_final'] for r in rows)
    unnamed = sorted({r['leadsheet'] for r in rows
                      if r['leadsheet'].strip() and not r['leadsheet_name']})
    print(f"OK   {name}: {len(rows)} accounts, {len(funds)} funds, {len(ajes)} JE lines, "
          f"{len(groupings)} named L/S groupings (CY final nets to {cy:,.2f})")
    if unnamed:
        print(f"WARN unnamed L/S codes (no MP match): {', '.join(unnamed)}")
    print(f"     -> {xlsx_path}")
    print(f"     -> {json_path}")


if __name__ == '__main__':
    main()
