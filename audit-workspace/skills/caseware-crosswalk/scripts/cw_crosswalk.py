"""Rebuild the CaseWare Document Manager index from a Working Papers folder copy.

CaseWare Working Papers folders are a flat pile of files on disk; the pretty,
indexed Document Manager tree lives in the engagement's <ClientFile>SH.dbf
(FoxPro DBF + .fpt memo file), which travels with any full copy of the folder.
Each record holds:
  SCHEDULE  - the workpaper index/reference shown in CaseWare (e.g. "IC-2")
  SCH_DESC  - the display name shown in CaseWare
  TYPE      - H = folder/header row, A = automatic doc (TB/leadsheets, live in
              the CaseWare database, no disk file), everything else = document
  MEMO      - for external documents, the PHYSICAL FILENAME on disk (only when
              it is a valid filename; automatic docs put descriptions here)
  DGN       - hierarchy code (first char - '!' = depth of the tree node)
  A1..A8 /  - sign-off initials and dates
  AD1..AD8
  deleted   - DBF delete flag / DELETED field = doc sits in the recycle bin

Usage:
  python cw_crosswalk.py "<engagement folder>" [--include-deleted] [--out DIR]
  python cw_crosswalk.py --all "<parent folder>" [--include-deleted]

Writes _crosswalk.xlsx (human) and _crosswalk.json (machine) into the top of
each engagement folder (or into --out if given). Falls back to _crosswalk.csv
if openpyxl is unavailable.
"""
import argparse
import csv
import datetime
import glob
import json
import os
import struct
import sys

COLS = ['index', 'caseware_name', 'folder', 'disk_file', 'on_disk',
        'doc_type', 'signoffs', 'modified', 'deleted']
HEADERS = ['Index', 'CaseWare Name', 'Folder', 'File on Disk', 'On Disk?',
           'Doc Type', 'Sign-offs', 'Modified', 'Deleted']


def read_dbf(path, memo_path=None):
    with open(path, 'rb') as f:
        hdr = f.read(32)
        nrec, = struct.unpack('<I', hdr[4:8])
        hdrlen, = struct.unpack('<H', hdr[8:10])
        reclen, = struct.unpack('<H', hdr[10:12])
        fields = []
        f.seek(32)
        while True:
            fd = f.read(32)
            if not fd or fd[0] == 0x0D:
                break
            name = fd[0:11].split(b'\x00')[0].decode('latin-1')
            fields.append((name, chr(fd[11]), fd[16]))
        memo = None
        blocksize = 512
        if memo_path and os.path.exists(memo_path):
            with open(memo_path, 'rb') as mf:
                memo = mf.read()
            blocksize = struct.unpack('>H', memo[6:8])[0] or 512
        f.seek(hdrlen)
        recs = []
        for _ in range(nrec):
            raw = f.read(reclen)
            if len(raw) < reclen:
                break
            rec = {'_deleted': raw[0:1] == b'*'}
            off = 1
            for name, ftype, flen in fields:
                chunk = raw[off:off + flen]
                off += flen
                if ftype == 'M':
                    val = ''
                    try:
                        blk = int(chunk.decode('latin-1').strip() or 0)
                    except ValueError:
                        blk = 0
                    if blk and memo:
                        pos = blk * blocksize
                        if pos + 8 <= len(memo):
                            _, mlen = struct.unpack('>II', memo[pos:pos + 8])
                            val = memo[pos + 8:pos + 8 + mlen].decode('latin-1', 'replace')
                    rec[name] = val.strip()
                else:
                    rec[name] = chunk.decode('latin-1', 'replace').strip()
            recs.append(rec)
        return recs


def find_sh_table(folder):
    hits = glob.glob(os.path.join(folder, '*SH.dbf')) + \
        glob.glob(os.path.join(folder, '*SH.DBF'))
    hits = [p for p in hits if os.sep + '_Sync' + os.sep not in p]
    return hits[0] if hits else None


def fmt_date(d):
    return f"{d[0:4]}-{d[4:6]}-{d[6:8]}" if d and len(d) >= 8 and d != '18991230' else ''


def build_rows(folder, include_deleted=False):
    sh = find_sh_table(folder)
    if not sh:
        return None, None
    fpt = None
    for ext in ('.fpt', '.FPT'):
        cand = os.path.splitext(sh)[0] + ext
        if os.path.exists(cand):
            fpt = cand
            break
    recs = read_dbf(sh, fpt)
    disk = {n.lower() for n in os.listdir(folder)}
    rows = []
    stack = []  # (depth, header name)
    for r in recs:
        deleted = r['_deleted'] or r.get('DELETED') == 'T'
        dgn = r.get('DGN', '')
        depth = (ord(dgn[0]) - ord('!')) if dgn else 0
        if r.get('TYPE') == 'H':
            if not deleted:
                stack = [s for s in stack if s[0] < depth]
                stack.append((depth, r.get('SCH_DESC', '')))
            continue
        if deleted and not include_deleted:
            continue
        memo_file = r.get('MEMO', '')
        # MEMO holds the physical filename for external docs; skip multi-line
        # notes and descriptions that can't be Windows filenames (e.g. the
        # "Leadsheet: Net Assets" placeholders on automatic docs)
        phys = memo_file
        if any(c in memo_file for c in '\r\n<>:"/\\|?*'):
            phys = ''
        on_disk = ''
        if phys:
            on_disk = 'yes' if phys.lower() in disk else 'MISSING'
        doctype = r.get('TYPE', '')
        kind = {'A': 'automatic (in CaseWare db)', 'M': 'manual'}.get(doctype, 'external file')
        signoffs = '; '.join(
            f"{r.get(f'A{i}', '')} {fmt_date(r.get(f'AD{i}', ''))}".strip()
            for i in range(1, 9) if r.get(f'A{i}'))
        rows.append({
            'index': r.get('SCHEDULE', ''),
            'caseware_name': r.get('SCH_DESC', ''),
            'folder': ' / '.join(s[1] for s in stack if s[0] < depth),
            'disk_file': phys,
            'on_disk': on_disk,
            'doc_type': kind,
            'signoffs': signoffs,
            'modified': fmt_date(r.get('CWDATE', '')),
            'deleted': 'yes' if deleted else '',
        })
    return rows, os.path.basename(sh)


def write_xlsx(rows, path, title):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Crosswalk'
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='44546A')
    for r in rows:
        ws.append([r[k] for k in COLS])
    for row in ws.iter_rows(min_row=2):
        if row[4].value == 'MISSING':
            for c in row:
                c.fill = PatternFill('solid', fgColor='FFF2CC')
    widths = [14, 55, 38, 55, 9, 24, 22, 11, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(path)


def process(folder, include_deleted=False, out_dir=None):
    folder = os.path.abspath(folder)
    name = os.path.basename(folder.rstrip('\\/')).replace(' (Sync)', '')
    rows, sh_name = build_rows(folder, include_deleted)
    if rows is None:
        print(f"SKIP {name}: no *SH.dbf found (need a full CaseWare folder copy)")
        return False
    out = out_dir or folder
    os.makedirs(out, exist_ok=True)
    meta = {
        'engagement': name,
        'generated': datetime.datetime.now().isoformat(timespec='seconds'),
        'source_table': sh_name,
        'note': ('Crosswalk of the CaseWare Document Manager index to the physical '
                 'files in this folder. disk_file is the real filename; '
                 'doc_type "automatic" docs live inside the CaseWare database '
                 'and have no disk file.'),
        'documents': rows,
    }
    json_path = os.path.join(out, '_crosswalk.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(meta, f, indent=1)
    try:
        xlsx_path = os.path.join(out, '_crosswalk.xlsx')
        write_xlsx(rows, xlsx_path, name)
        pretty = xlsx_path
    except ImportError:
        pretty = os.path.join(out, '_crosswalk.csv')
        with open(pretty, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=COLS)
            w.writeheader()
            w.writerows(rows)
    ext = [r for r in rows if r['disk_file'] and not r['deleted']]
    missing = [r for r in ext if r['on_disk'] == 'MISSING']
    print(f"OK   {name}: {len([r for r in rows if not r['deleted']])} documents, "
          f"{len(ext)} external files, {len(missing)} missing on disk")
    print(f"     -> {pretty}")
    print(f"     -> {json_path}")
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('folder', help='engagement folder, or parent folder with --all')
    ap.add_argument('--all', action='store_true',
                    help='process every CaseWare folder directly inside')
    ap.add_argument('--out', default=None,
                    help='write outputs here instead of into the engagement folder')
    ap.add_argument('--include-deleted', action='store_true',
                    help='include recycle-bin documents')
    args = ap.parse_args()
    if args.all:
        found = 0
        for d in sorted(os.listdir(args.folder)):
            full = os.path.join(args.folder, d)
            if os.path.isdir(full) and find_sh_table(full):
                process(full, args.include_deleted, args.out)
                found += 1
        if not found:
            print("No CaseWare folders (with *SH.dbf) found inside", args.folder)
            sys.exit(1)
    else:
        if not process(args.folder, args.include_deleted, args.out):
            sys.exit(1)


if __name__ == '__main__':
    main()
