"""Render the merged findings JSON into the deliverable Excel report.

Input schema (see references/output-spec.md for the full deliverable spec):
{
  "meta": {"client": "...", "fs_date": "...", "framework": "commercial|nonprofit|govt|ebp",
           "limitations": ["..."]},
  "findings": [{"id","step","category","severity","location","description","recommendation"}],
  "procedures": [{"tab": "Proof Review", "items":
                   [{"procedure","result": "Pass|Fail|Flag","notes"}]}]
}
Findings tabs are derived from each finding's "step"; procedure tabs render as listed.

Usage: python report.py findings.json -o "Client FS Review.xlsx"
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from fslib import load_json

SEV_ORDER = {"Critical": 0, "Significant": 1, "Moderate": 2, "Minor": 3}
STEP_TABS = [
    ("proof", "Proof Review"),
    ("report-language", "Report Language"),
    ("math", "Math Check"),
    ("xref", "Cross-Reference"),
    ("disclosure", "Disclosure Review"),
    ("industry", "Industry Review"),
    ("supplemental", "Supplemental Schedules"),
    ("final", "Final Proof Checklist"),
]
HDR = Font(bold=True, underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def _header_block(ws, meta, tabname):
    ws["A1"] = meta.get("client", "")
    ws["A2"] = tabname
    ws["A3"] = meta.get("fs_date", "")
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = Font(bold=(r == 1))


def _write_table(ws, start_row, headers, rows, widths, wrap_cols):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.font = HDR
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    r = start_row + 1
    for row in rows:
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.alignment = WRAP if ci in wrap_cols else TOP
        r += 1
    return r


def build(data, out_path):
    meta = data.get("meta", {})
    findings = sorted(data.get("findings", []),
                      key=lambda f: (SEV_ORDER.get(f.get("severity"), 9),
                                     f.get("id", "")))
    wb = Workbook()

    # Tab 1 — Executive Summary
    ws = wb.active
    ws.title = "Executive Summary"
    _header_block(ws, meta, "Executive Summary")
    r = 5
    ws.cell(row=r, column=1, value="Findings by severity").font = HDR
    r += 1
    for sev in ("Critical", "Significant", "Moderate", "Minor"):
        n = sum(1 for f in findings if f.get("severity") == sev)
        ws.cell(row=r, column=1, value=sev)
        ws.cell(row=r, column=2, value=n)
        r += 1
    for lim in meta.get("limitations", []):
        r += 1
        ws.cell(row=r, column=1, value="Limitation")
        ws.cell(row=r, column=2, value=lim)  # no wrap — let it overflow
    r += 2
    rows = [[f.get("id"), f.get("step"), f.get("category"), f.get("severity"),
             f.get("location"), f.get("description"), f.get("recommendation")]
            for f in findings]
    _write_table(ws, r,
                 ["Finding ID", "Step", "Category", "Severity", "Location",
                  "Description", "Recommended Correction"],
                 rows, [12, 16, 22, 12, 26, 70, 55], wrap_cols={6, 7})

    # Findings detail tabs by step
    for step_key, tabname in STEP_TABS:
        step_findings = [f for f in findings if f.get("step") == step_key]
        procs = [p for p in data.get("procedures", []) if p.get("tab") == tabname]
        if not step_findings and not procs:
            continue
        ws = wb.create_sheet(tabname[:31])
        _header_block(ws, meta, tabname)
        r = 5
        if step_findings:
            ws.cell(row=r, column=1, value="Findings").font = HDR
            r += 1
            rows = [[f.get("id"), f.get("category"), f.get("severity"),
                     f.get("location"), f.get("description"), f.get("recommendation")]
                    for f in step_findings]
            r = _write_table(ws, r,
                             ["Finding ID", "Category", "Severity", "Location",
                              "Description", "Recommended Correction"],
                             rows, [12, 22, 12, 26, 70, 55], wrap_cols={5, 6}) + 1
        for p in procs:
            rows = [[i.get("procedure"), i.get("result"), i.get("notes", "")]
                    for i in p.get("items", [])]
            ws.cell(row=r, column=1, value="Procedures").font = HDR
            r = _write_table(ws, r + 1, ["Procedure", "Result", "Notes"],
                             rows, [70, 10, 55], wrap_cols={1, 3}) + 1

    wb.save(out_path)
    # integrity check — reopen to prove the file isn't corrupt
    from openpyxl import load_workbook
    load_workbook(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("findings")
    ap.add_argument("-o", "--out", default="fs-review-report.xlsx")
    args = ap.parse_args()
    data = load_json(args.findings)
    out = build(data, args.out)
    n = len(data.get("findings", []))
    print(f"{n} findings -> {out} (reopened cleanly)")


if __name__ == "__main__":
    main()
