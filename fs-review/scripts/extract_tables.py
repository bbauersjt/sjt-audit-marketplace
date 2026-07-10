"""Extract every financial-statement table from a PDF (or xlsx) into statements.json.

Column mapping is done geometrically: money tokens are right-aligned, so their right
edges cluster into columns. Every row's values are snapped to those columns, which is
what makes multi-column (consolidating/combining/WIP) statements reliable without the
source Excel. Self-checks emit warnings when tokens don't snap cleanly — those pages
should be verified visually instead of trusted.

Usage:
    python extract_tables.py <input.pdf|input.xlsx> [-o statements.json] [--pages 1-99]
Output schema:
    {"source": ..., "tables": [{"id","page","title","columns":[...],
      "rows":[{"label","values":[float|None,...],"raw":[...]}], "warnings":[...]}]}
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from fslib import is_money_token, parse_money, save_json

LINE_Y_TOL = 3.0      # pts: words within this vertical distance are one line
COL_X_TOL = 6.0       # pts: right-edge cluster tolerance for a column
GAP_NEW_TABLE = 28.0  # pts: vertical gap that splits two tables on one page


def _group_lines(words):
    """Group pdfplumber words into lines by y midpoint."""
    lines = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        ymid = (w["top"] + w["bottom"]) / 2
        for line in lines:
            if abs(line["y"] - ymid) <= LINE_Y_TOL:
                line["words"].append(w)
                break
        else:
            lines.append({"y": ymid, "words": [w]})
    for line in lines:
        line["words"].sort(key=lambda w: w["x0"])
    return sorted(lines, key=lambda l: l["y"])


def _merge_money_tokens(words):
    """Rejoin money values pdfplumber split.

    Two split classes: symbol attachment ('$' '1,234' / '(1,234' ')') at up to a
    small-word-space gap, and kerning splits INSIDE a rendered number ('9' '88,553',
    '2' ',724,357' — common in Word-exported PDFs) at near-zero gaps. The kerning
    threshold must stay below real word-space width (~2.5pt at 8-9pt fonts) so two
    genuinely separate tokens never fuse."""
    merged = []
    for w in words:
        t = w["text"]
        if merged:
            prev = merged[-1]
            gap = w["x0"] - prev["x1"]
            joined = prev["text"] + t
            symbol_attach = prev["text"] in ("$", "(") or t == ")"
            kerning_split = (is_money_token(joined)
                             and re.fullmatch(r"[$(]?[\d,.]+", prev["text"] or " ")
                             and re.fullmatch(r"[\d,.]+\)?", t or " "))
            if (symbol_attach and gap < 4.5) or (kerning_split and gap < 2.0):
                prev["text"] = joined
                prev["x1"] = w["x1"]
                continue
        merged.append(dict(w))
    return merged


STRONG_MONEY = re.compile(r"[,$()]|\.\d")


def _cluster_columns(lines):
    """Cluster right edges of money tokens on a set of lines into column x positions.

    Only strongly money-like tokens (comma/$/parens/decimals) can FOUND a column —
    bare integers (years, note numbers, job numbers) and dash placeholders can only
    snap to columns founded by real values."""
    edges = []
    for line in lines:
        for w in line["words"]:
            if is_money_token(w["text"]) and STRONG_MONEY.search(w["text"]):
                edges.append(w["x1"])
    edges.sort()
    cols = []
    for x in edges:
        if cols and x - cols[-1][-1] <= COL_X_TOL:
            cols[-1].append(x)
        else:
            cols.append([x])
    # a column must appear on >=2 lines to count (kills stray footnote numbers)
    return [sum(c) / len(c) for c in cols if len(c) >= 2]


def _column_headers(header_lines_words, col_xs):
    """Assign header words to columns geometrically (right-aligned headers share the
    column's right edge; wrapped multi-line headers accumulate top-to-bottom)."""
    headers = [[] for _ in col_xs]
    for words in header_lines_words:
        for w in words:
            best, bestd = None, None
            for i, cx in enumerate(col_xs):
                center = (w["x0"] + w["x1"]) / 2
                d = min(abs(w["x1"] - cx), abs(center - cx) * 0.6)
                if bestd is None or d < bestd:
                    best, bestd = i, d
            if bestd is not None and bestd <= 30:
                headers[best].append(w["text"])
    return [" ".join(h) for h in headers]


def _snap(word, col_xs):
    """Index of the column this money token belongs to, or None."""
    best, bestd = None, None
    for i, cx in enumerate(col_xs):
        d = abs(word["x1"] - cx)
        if bestd is None or d < bestd:
            best, bestd = i, d
    if bestd is not None and bestd <= COL_X_TOL + 2:
        return best
    return None


def _extract_page(page, pageno):
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    if not words:
        return []
    lines = _group_lines(words)
    for line in lines:
        line["words"] = _merge_money_tokens(line["words"])
        line["has_money"] = any(is_money_token(w["text"]) for w in line["words"])

    # split page into blocks at large vertical gaps between money-bearing regions
    blocks, cur = [], []
    last_y = None
    for line in lines:
        if last_y is not None and line["y"] - last_y > GAP_NEW_TABLE and any(
                l["has_money"] for l in cur):
            blocks.append(cur)
            cur = []
        cur.append(line)
        last_y = line["y"]
    if cur:
        blocks.append(cur)

    # remerge adjacent blocks that share the same column grid — Word section spacing
    # splits a single statement into blocks, which would orphan its totals from
    # their components
    def _sig(block):
        return _cluster_columns([l for l in block if l["has_money"]])

    def _same_grid(a, b):
        if not a or not b:
            return False
        small, large = (a, b) if len(a) <= len(b) else (b, a)
        return all(any(abs(x - y) <= COL_X_TOL + 2 for y in large) for x in small)

    merged_blocks = []
    for block in blocks:
        if merged_blocks and _same_grid(_sig(merged_blocks[-1]), _sig(block)):
            merged_blocks[-1].extend(block)
        else:
            merged_blocks.append(block)
    blocks = merged_blocks

    tables = []
    for bi, block in enumerate(blocks):
        money_lines = [l for l in block if l["has_money"]]
        if len(money_lines) < 2:
            continue
        col_xs = _cluster_columns(money_lines)
        if not col_xs:
            continue
        warnings = []
        rows = []
        header_lines = []
        header_words = []
        for line in block:
            toks = line["words"]
            has_strong = any(is_money_token(w["text"]) and STRONG_MONEY.search(w["text"])
                             for w in toks)
            if not has_strong:
                # a line whose values are all under 1,000 has no comma/$/() token —
                # still a data row if a (non-year) money token snaps to the grid
                snaps = any(
                    is_money_token(w["text"])
                    and not re.fullmatch(r"(19|20)\d{2}", w["text"])
                    and _snap(w, col_xs) is not None
                    for w in toks)
                if not snaps:
                    text = " ".join(w["text"] for w in toks)
                    if rows:
                        rows.append({"label": text, "values": [None] * len(col_xs),
                                     "raw": [""] * len(col_xs)})
                    else:
                        header_lines.append(text)
                        header_words.append(toks)
                    continue
            values = [None] * len(col_xs)
            raw = [""] * len(col_xs)
            label_parts = []
            offgrid = []
            for w in toks:
                if is_money_token(w["text"]):
                    ci = _snap(w, col_xs)
                    if ci is None:
                        if re.search(r"[,$()]", w["text"]):
                            offgrid.append(w["text"])
                        label_parts.append(w["text"])
                    elif values[ci] is not None:
                        warnings.append(
                            f"p{pageno}: two values landed in column {ci+1} on line "
                            f"'{' '.join(t['text'] for t in toks)[:60]}' — verify visually")
                    else:
                        values[ci] = parse_money(w["text"])
                        raw[ci] = w["text"]
                else:
                    label_parts.append(w["text"])
            if not any(v is not None for v in values):
                # nothing snapped: a prose line with inline amounts ("a $15,000,000
                # line of credit"), not a misparsed table row — keep as label only
                rows.append({"label": " ".join(w["text"] for w in toks).strip(),
                             "values": values, "raw": raw})
                continue
            # warn only for genuinely tabular lines (some values snapped, one didn't)
            for tok in offgrid:
                warnings.append(
                    f"p{pageno}: token '{tok}' did not snap to any column on a line "
                    f"with snapped values — verify this line visually")
            rows.append({"label": " ".join(label_parts).strip(), "values": values,
                         "raw": raw})
        # merge wrapped labels: a value row whose label starts lowercase (or is
        # empty) continues the label-only line above it ("Net cash used by noncapital"
        # / "financing activities   (59,486,202)")
        merged_rows = []
        for row in rows:
            has_vals = any(v is not None for v in row["values"])
            prev = merged_rows[-1] if merged_rows else None
            if (has_vals and prev is not None
                    and not any(v is not None for v in prev["values"])
                    and prev["label"]
                    and (not row["label"] or row["label"][:1].islower())):
                row["label"] = (prev["label"] + " " + row["label"]).strip()
                merged_rows[-1] = row
            else:
                merged_rows.append(row)
        rows = merged_rows

        # column headers: geometric assignment of header words to columns
        colheads = _column_headers(header_words, col_xs)
        columns = [ch if ch.strip() else f"col{i+1}"
                   for i, ch in enumerate(colheads)]
        yearish = [h for h in header_lines if re.search(r"\b(19|20)\d{2}\b", h)]
        title = header_lines[0].strip() if header_lines else ""
        tables.append({
            "id": f"p{pageno:03d}t{bi + 1}",
            "page": pageno,
            "title": title,
            "header_lines": header_lines,
            "columns": columns,
            "column_hint": yearish[-1] if yearish else "",
            "rows": rows,
            "warnings": warnings,
        })
    return tables


def extract_pdf(path, pages=None):
    import pdfplumber
    tables = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            if pages and i not in pages:
                continue
            tables.extend(_extract_page(page, i))
    return tables


def extract_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    tables = []
    for si, ws in enumerate(wb.worksheets, start=1):
        rows = []
        maxc = ws.max_column
        for r in ws.iter_rows(values_only=True):
            cells = list(r) + [None] * (maxc - len(r))
            label = ""
            values, raw = [], []
            for c in cells:
                if isinstance(c, (int, float)):
                    values.append(float(c))
                    raw.append(str(c))
                elif c is None:
                    values.append(None)
                    raw.append("")
                else:
                    v = parse_money(str(c))
                    if v is not None and not label:
                        values.append(v)
                        raw.append(str(c))
                    else:
                        label = (label + " " + str(c)).strip()
                        values.append(None)
                        raw.append("")
            if label or any(v is not None for v in values):
                rows.append({"label": label, "values": values, "raw": raw})
        if rows:
            ncols = max(len(r["values"]) for r in rows)
            for r in rows:
                r["values"] += [None] * (ncols - len(r["values"]))
                r["raw"] += [""] * (ncols - len(r["raw"]))
            tables.append({"id": f"sheet{si}", "page": si, "title": ws.title,
                           "header_lines": [], "columns": [f"col{i+1}" for i in range(ncols)],
                           "column_hint": "", "rows": rows, "warnings": []})
    return tables


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("-o", "--out", default="statements.json")
    ap.add_argument("--pages", help="e.g. 3-40 (PDF only)")
    args = ap.parse_args()
    pages = None
    if args.pages:
        a, _, b = args.pages.partition("-")
        pages = set(range(int(a), int(b or a) + 1))
    p = Path(args.input)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        tables = extract_xlsx(p)
    else:
        tables = extract_pdf(p, pages)
    nwarn = sum(len(t["warnings"]) for t in tables)
    out = {"source": str(p), "tables": tables}
    save_json(out, args.out)
    print(f"{len(tables)} tables -> {args.out}; {nwarn} extraction warnings"
          + (" (verify flagged lines visually before treating diffs as findings)" if nwarn else ""))
    for t in tables:
        for w in t["warnings"]:
            print("  WARN", t["id"], w)


if __name__ == "__main__":
    main()
