"""
tb_io.py — IO helpers for the trial-balance-prep skill.

This module does NOT make judgment calls. It only:
  - reads trial balances and grouped/summary TBs into normalized rows,
  - reads a 4-digit grouping index workbook,
  - detects whether a set of grouping codes is legacy or new format,
  - reads the classification vocabulary (default-classes.xlsx),
  - infers a fund identifier from an account-number string (best-effort),
  - writes the grouping-conversion review workbook (decorated, for the user),
  - writes sterile CCH import files (Basic / Grouped / Fund tiers).

The semantic work — matching an old code to a new 4-digit code, picking a
Classification abbreviation, deciding a fund's identity — is performed by the
skill at run time, guided by SKILL.md. That context-driven judgment
deliberately does not live in code; these helpers only move data and enforce
the mechanical output rules (CCH constraints, number formats).

Dependency: openpyxl only. Run from bash via the skill's scripts/ path.
"""

import csv
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# --------------------------------------------------------------------------
# Code normalization
# --------------------------------------------------------------------------

def norm_code(value):
    """Normalize a grouping code for comparison.

    '  B. 1 ' -> 'B.1' ; ' N ' -> 'N' ; '10' -> '10' ; '1000' -> '1000'.
    Collapses whitespace around an internal dot so sub-codes compare cleanly.
    Returns '' for blank/None.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = re.sub(r"\s*\.\s*", ".", s)   # 'B. 1' -> 'B.1'
    s = re.sub(r"\s+", " ", s)
    return s


LEGACY_RE = re.compile(r"^[A-Za-z]{1,3}(\.\d+)?$|^\d{1,2}(\.\d+)?$")
NEW_RE = re.compile(r"^\d{4}$")


def detect_format(codes):
    """Classify a collection of grouping codes.

    Returns one of: 'new', 'legacy', 'mixed', 'empty'.
    """
    seen = [norm_code(c) for c in codes]
    seen = [c for c in seen if c]
    if not seen:
        return "empty"
    is_new = all(NEW_RE.match(c) for c in seen)
    is_legacy = all(LEGACY_RE.match(c) for c in seen)
    if is_new:
        return "new"
    if is_legacy:
        return "legacy"
    return "mixed"


def classify_codes(codes):
    """Per-code classification, for diagnosing a 'mixed' column.

    Returns dict {code: 'new'|'legacy'|'unknown'}.
    """
    out = {}
    for c in codes:
        n = norm_code(c)
        if not n:
            continue
        if NEW_RE.match(n):
            out[n] = "new"
        elif LEGACY_RE.match(n):
            out[n] = "legacy"
        else:
            out[n] = "unknown"
    return out


# --------------------------------------------------------------------------
# Reading tabular files
# --------------------------------------------------------------------------

def _rows_from_file(path):
    """Yield raw rows (lists of cell values) from an xlsx or csv file."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    elif ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.reader(fh, delimiter=delim):
                yield row
    else:
        raise ValueError("Unsupported file type: %s" % ext)


def _to_number(value):
    """Parse a balance cell. Handles parentheses-negatives, $, commas."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").strip()
    if s in ("-", "--"):
        return 0.0
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


# Header keyword -> logical column
_HEADER_HINTS = {
    "account": "account", "acct": "account", "number": "account",
    "name": "name", "descr": "name", "description": "name", "title": "name",
    "l/s": "code", "ls": "code", "lead": "code", "leadsheet": "code",
    "group": "code", "grouping": "code",
    "balance": "balance", "amount": "balance", "current": "balance",
    "cy": "balance", "prelim": "balance",
}


def _find_header(rows, max_scan=15):
    """Locate the header row; return (header_index, {logical: col_index})."""
    for i, row in enumerate(rows[:max_scan]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        mapping = {}
        for ci, cell in enumerate(cells):
            for hint, logical in _HEADER_HINTS.items():
                if hint in cell and logical not in mapping:
                    mapping[logical] = ci
        # a real header has at least a name-ish and a balance-ish column
        if "name" in mapping and "balance" in mapping:
            return i, mapping
    return None, {}


def load_tb(path):
    """Load an account-level trial balance.

    Returns list of dicts: {row, account, name, code, balance}.
    'code' is the raw grouping/leadsheet code as found (not normalized).
    """
    rows = list(_rows_from_file(path))
    h_idx, cols = _find_header(rows)
    if h_idx is None:
        raise ValueError(
            "Could not locate a header row in %s — inspect the file manually." % path
        )
    out = []
    for r in range(h_idx + 1, len(rows)):
        row = rows[r]

        def cell(logical):
            ci = cols.get(logical)
            if ci is None or ci >= len(row):
                return None
            return row[ci]

        name = cell("name")
        if name is None or str(name).strip() == "":
            continue
        bal = _to_number(cell("balance"))
        out.append({
            "row": r + 1,
            "account": (str(cell("account")).strip()
                        if cell("account") is not None else ""),
            "name": str(name).strip(),
            "code": (str(cell("code")).strip()
                     if cell("code") is not None else ""),
            "balance": bal if bal is not None else 0.0,
        })
    return out


# A grouped/summary TB row looks like:  "N  Investments"  /  "B. 1  Participant loans"
_LEGEND_RE = re.compile(r"^\s*([A-Za-z0-9]{1,3}(?:\.\s?\d+)?)\s{2,}(.+?)\s*$")
# fallback: single token with no label, e.g. "M"
_LEGEND_BARE_RE = re.compile(r"^\s*([A-Za-z0-9]{1,3}(?:\.\s?\d+)?)\s*$")


def load_legend(path):
    """Parse a grouped / summary TB into {old_code: label}.

    Best-effort. The summary's first text column carries '<code>  <label>'
    strings. Rows that don't parse as a code+label are skipped. The skill
    should always eyeball the file to confirm / supplement this.
    """
    rows = list(_rows_from_file(path))
    legend = {}
    for row in rows:
        if not row:
            continue
        first = row[0]
        if first is None:
            continue
        text = str(first).strip()
        if not text:
            continue
        m = _LEGEND_RE.match(text)
        if m:
            code = norm_code(m.group(1))
            label = m.group(2).strip()
            if code and code not in legend:
                legend[code] = label
            continue
        m = _LEGEND_BARE_RE.match(text)
        if m:
            code = norm_code(m.group(1))
            # only treat as a code if it looks like a legacy code
            if code and LEGACY_RE.match(code) and code not in legend:
                legend[code] = ""
    return legend


# --------------------------------------------------------------------------
# Reading a grouping index workbook
# --------------------------------------------------------------------------

def load_index(path, sheet=None):
    """Load a 4-digit grouping index.

    Returns list of dicts: {index, group_name, class, statement, ...}.
    Reads the first sheet unless `sheet` is given. Tolerates the header
    sitting a few rows down (title/subtitle rows above it).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    h_idx, hdr = None, []
    for i, row in enumerate(rows[:12]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "index" in cells and any("group" in c for c in cells):
            h_idx, hdr = i, cells
            break
    if h_idx is None:
        raise ValueError("No 'Index' header found in %s" % path)

    def col(*keys):
        for ci, c in enumerate(hdr):
            if any(k in c for k in keys):
                return ci
        return None

    ci_idx = col("index")
    ci_grp = col("group")
    ci_cls = col("class")
    ci_stmt = col("statement", "financial")

    out = []
    for r in range(h_idx + 1, len(rows)):
        row = rows[r]
        if ci_idx is None or ci_idx >= len(row) or row[ci_idx] is None:
            continue
        idx = str(row[ci_idx]).replace("\n", "").strip()
        if not NEW_RE.match(idx):
            continue
        grp = ""
        if ci_grp is not None and ci_grp < len(row) and row[ci_grp] is not None:
            grp = str(row[ci_grp]).replace("\n", "").strip()
        rec = {"index": idx, "group_name": grp, "class": "", "statement": ""}
        if ci_cls is not None and ci_cls < len(row) and row[ci_cls] is not None:
            rec["class"] = str(row[ci_cls]).strip()
        if ci_stmt is not None and ci_stmt < len(row) and row[ci_stmt] is not None:
            rec["statement"] = str(row[ci_stmt]).strip()
        out.append(rec)
    return out


# --------------------------------------------------------------------------
# Writing the conversion output workbook
# --------------------------------------------------------------------------

_NUMFMT = '#,##0.00;(#,##0.00);"-"'
_NAVY = PatternFill("solid", fgColor="1F3864")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_sheet(ws, headers, rows, numeric_cols=()):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
    for ri, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=c, value=val)
            cell.border = _BORDER
            if c in numeric_cols:
                cell.number_format = _NUMFMT
    ws.freeze_panes = "A2"
    for c in range(1, len(headers) + 1):
        width = max(12, min(48, max(
            [len(str(headers[c - 1]))] +
            [len(str(r[c - 1])) for r in rows if c - 1 < len(r)] + [0]
        ) + 2))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = width


def write_output(out_path, code_map, converted_rows, log_rows):
    """Write the 3-sheet conversion workbook.

    code_map: list of dicts with keys old_code, old_meaning, new_index,
              new_group, klass, source, confidence, notes
    converted_rows: list of dicts with keys account, name, old_code,
              new_index, new_group, balance
    log_rows: list of dicts with keys account, name, message
    """
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Code Map"
    _write_sheet(
        ws,
        ["Old Code", "Old Meaning", "New Index", "New Group Name",
         "Class", "Match Source", "Confidence", "Notes"],
        [[m.get("old_code", ""), m.get("old_meaning", ""),
          m.get("new_index", ""), m.get("new_group", ""),
          m.get("klass", ""), m.get("source", ""),
          m.get("confidence", ""), m.get("notes", "")]
         for m in code_map],
    )

    ws = wb.create_sheet("Converted TB")
    _write_sheet(
        ws,
        ["Account #", "Account Name", "Old Code", "New Index",
         "New Group Name", "Balance"],
        [[r.get("account", ""), r.get("name", ""), r.get("old_code", ""),
          r.get("new_index", ""), r.get("new_group", ""),
          r.get("balance", 0.0)]
         for r in converted_rows],
        numeric_cols=(6,),
    )
    total = sum(float(r.get("balance", 0) or 0) for r in converted_rows)
    tr = len(converted_rows) + 2
    ws.cell(row=tr, column=2, value="TOTAL").font = Font(bold=True)
    tc = ws.cell(row=tr, column=6, value=total)
    tc.number_format = _NUMFMT
    tc.font = Font(bold=True)

    ws = wb.create_sheet("Conversion Log")
    _write_sheet(
        ws,
        ["Account #", "Account Name", "Message"],
        [[r.get("account", ""), r.get("name", ""), r.get("message", "")]
         for r in log_rows],
    )

    wb.save(out_path)
    return {"code_map": len(code_map), "rows": len(converted_rows),
            "log": len(log_rows), "tb_total": total}


# --------------------------------------------------------------------------
# Classification vocabulary (default-classes.xlsx)
# --------------------------------------------------------------------------

# The firm's 10-way import classification. Order = balance-sheet then
# income-statement. This is the vocabulary the CCH "Classification" column
# must draw from. It is FINER than the coarse "Class" carried on the grouping
# index (Asset/Liability/Equity/Revenue/Expense/Other), so mapping from the
# index class to one of these requires two judgment axes the index does not
# carry: current vs non-current, and operating vs cost-of-revenue vs other.
CLASS_ABBREVS = ["CA", "NA", "CL", "NL", "EQ", "REV", "COR", "OI", "OPX", "OE"]


def load_classes(path):
    """Load default-classes.xlsx -> list of {abbrev, description} in order.

    Header row is 'ABBREV | DESCRIPTION'. Returns the firm vocabulary so the
    skill validates every Classification it writes against the real list
    instead of inventing values.
    """
    rows = list(_rows_from_file(path))
    out = []
    started = False
    for row in rows:
        if not row or row[0] is None:
            continue
        a = str(row[0]).strip()
        d = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not started:
            if a.upper() in ("ABBREV", "ABBREVIATION", "CLASS"):
                started = True
            continue
        if a:
            out.append({"abbrev": a, "description": d})
    return out


# Coarse index Class -> the abbreviations it can legitimately resolve to.
# The FIRST entry is the safe default when no finer signal is available; the
# skill upgrades to a non-current / cost / other choice using group-name and
# account-nature hints (see guess_class).
_CLASS_FROM_COARSE = {
    "asset": ["CA", "NA"],
    "liability": ["CL", "NL"],
    "equity": ["EQ"],
    "net assets": ["EQ"],           # EBP / nonprofit wording
    "revenue": ["REV", "OI"],
    "addition": ["REV", "OI"],      # EBP additions
    "expense": ["OPX", "COR", "OE"],
    "deduction": ["OPX", "OE"],     # EBP deductions
    "other": ["OI", "OE"],
    "transfer": ["OPX", "REV"],     # governmental transfers; direction decides
}

# Keyword hints that push an Asset/Liability to the NON-current abbreviation.
# Kept precise on purpose: bare "property" would mis-flag "property taxes
# receivable", and a generic "deferred inflow/outflow" is current as often as
# not — so only pension/OPEB deferrals count here.
_NONCURRENT_HINTS = (
    "property, plant", "property plant", "plant and equipment", "ppe", "p,p&e",
    "capital asset", "buildings", "infrastructure", "machinery", "land",
    "construction in progress", "accumulated depreciation", "intangible",
    "goodwill", "right-of-use", "rou ", "long-term", "long term", "noncurrent",
    "non-current", "pension", "opeb", "bonds payable", "notes payable",
    "lease liability", "compensated absences - noncurrent",
)
# If present, force CURRENT regardless of any non-current hint above.
_CURRENT_OVERRIDE = ("current portion", "current maturit", "- current", "current)")
# Keyword hints for expense sub-type.
_COR_HINTS = ("cost of", "cost of goods", "cogs", "cost of revenue", "cost of sales")
_OTHER_EXP_HINTS = ("interest expense", "loss on", "impairment", "other expense")
_OTHER_INC_HINTS = ("gain on", "other income", "interest income", "dividend")


def guess_class(coarse, group_name="", account_name="", noncurrent=None):
    """Best-guess a 10-way abbreviation from the coarse index Class + hints.

    Returns (abbrev, confidence, note). This is a HINT, not an answer — the
    skill confirms low/medium calls. `noncurrent`, when passed explicitly
    (True/False), overrides keyword inference.
    """
    c = (coarse or "").strip().lower()
    opts = _CLASS_FROM_COARSE.get(c)
    if not opts:
        return ("", "low", "coarse class %r not recognized" % coarse)

    text = ("%s %s" % (group_name, account_name)).lower()

    if c in ("asset", "liability"):
        nc = noncurrent
        if nc is None:
            if any(h in text for h in _CURRENT_OVERRIDE):
                nc = False
            else:
                nc = any(h in text for h in _NONCURRENT_HINTS)
        abbr = opts[1] if nc else opts[0]
        conf = "high" if noncurrent is not None else ("medium" if nc else "high")
        note = "" if not nc else "classified non-current on keyword/nature — confirm"
        return (abbr, conf, note)

    if c in ("expense", "deduction"):
        if any(h in text for h in _COR_HINTS):
            return ("COR", "medium", "looks like cost of revenue — confirm vs OPX")
        if any(h in text for h in _OTHER_EXP_HINTS):
            return ("OE", "medium", "looks like non-operating expense — confirm vs OPX")
        return ("OPX", "high", "")

    if c in ("revenue", "addition"):
        if any(h in text for h in _OTHER_INC_HINTS):
            return ("OI", "medium", "looks like other/non-operating income — confirm vs REV")
        return ("REV", "high", "")

    if c == "transfer":
        # governmental: transfers-in read as REV, transfers-out as OPX
        if "in" in text and "out" not in text:
            return ("REV", "medium", "transfer in — confirm direction")
        return ("OPX", "medium", "transfer out — confirm direction")

    return (opts[0], "high", "")


# --------------------------------------------------------------------------
# Fund inference from an account-number string
# --------------------------------------------------------------------------

# 100-10000  /  100:10000  -> fund is the leading segment
_FUND_PREFIX_RE = re.compile(r"^\s*(\d{2,4})\s*[-:.]\s*\d")
# 10000-100  -> fund is the trailing segment
_FUND_SUFFIX_RE = re.compile(r"\d\s*[-:.]\s*(\d{2,4})\s*$")


def infer_fund(account):
    """Guess a fund identifier from an account-number string.

    Returns (fund, pattern) where pattern is 'prefix' | 'suffix' | None.
    Only structural inference — a TB that carries fund NAMES, or a Caseware
    export whose funds live in the folder tree, must be handled by the skill
    (see SKILL.md, fund workflow). Never guess a fund silently for those.
    """
    if account is None:
        return (None, None)
    s = str(account).strip()
    m = _FUND_PREFIX_RE.match(s)
    if m:
        return (m.group(1), "prefix")
    m = _FUND_SUFFIX_RE.search(s)
    if m:
        return (m.group(1), "suffix")
    return (None, None)


def scan_fund_patterns(accounts):
    """Summarize fund inference across a whole TB.

    Returns {'pattern': dominant_pattern_or_None, 'funds': sorted_list,
    'coverage': fraction_of_accounts_with_a_fund}. The skill uses coverage +
    consistency to decide whether it is confident enough to proceed or must
    confirm/ask.
    """
    pats, funds, hit = {}, set(), 0
    accts = list(accounts)
    for a in accts:
        f, p = infer_fund(a)
        if f:
            hit += 1
            funds.add(f)
            pats[p] = pats.get(p, 0) + 1
    dominant = max(pats, key=pats.get) if pats else None
    cov = (hit / len(accts)) if accts else 0.0
    return {"pattern": dominant, "funds": sorted(funds), "coverage": round(cov, 3)}


# --------------------------------------------------------------------------
# Writing sterile CCH import files (Basic / Grouped / Fund tiers)
# --------------------------------------------------------------------------
#
# CCH import constraints baked in here (see references/cch-import-formats.md):
#   - NO freeze panes (hard reject).      - Header in row 1, no title rows.
#   - Single balance column.              - Debits positive / credits negative.
#   - Account number stays TEXT (leading zeros / dash prefixes survive).
#   - Accounting number format on balance columns.
# These writers are deliberately undecorated — no fills, no borders. The
# decorated three-sheet review workbook is write_output(), a different file
# for a different consumer.

_CCH_NUMFMT = '#,##0.00;(#,##0.00);"-"'

_CCH_HEADERS = {
    "basic": ["Account Number", "Account Name", "Account Balance"],
    "grouped": ["Account Number", "Account Name", "Account Balance",
                "Prior Year Balance", "Group Index", "Group Name",
                "Subgroup Index", "Subgroup Name", "Classification"],
    "fund": ["Account Number", "Account Name", "Account Balance",
             "Prior Year Balance", "Group Index", "Group Name",
             "Subgroup Index", "Subgroup Name", "Classification", "Fund Index"],
}

# Which output columns are numeric (accounting format) vs must stay text.
_CCH_NUMERIC = {"basic": (3,), "grouped": (3, 4), "fund": (3, 4)}
_CCH_TEXT = {
    "basic": (1,),
    "grouped": (1, 5, 7),          # account #, group index, subgroup index
    "fund": (1, 5, 7, 10),         # + fund index
}

# The key each header column pulls from a row dict.
_CCH_KEYS = {
    "Account Number": "account", "Account Name": "name",
    "Account Balance": "balance", "Prior Year Balance": "py_balance",
    "Group Index": "group_index", "Group Name": "group_name",
    "Subgroup Index": "subgroup_index", "Subgroup Name": "subgroup_name",
    "Classification": "classification", "Fund Index": "fund_index",
}


def write_cch_import(out_path, tier, rows, sheet_name="TB"):
    """Write a sterile CCH import workbook for one tier.

    tier: 'basic' | 'grouped' | 'fund'
    rows: list of dicts. Keys used per tier (missing keys -> blank):
          account, name, balance, py_balance, group_index, group_name,
          subgroup_index, subgroup_name, classification, fund_index
    Returns {'tier', 'rows', 'total', 'by_fund'}.
    """
    tier = tier.lower()
    if tier not in _CCH_HEADERS:
        raise ValueError("tier must be basic|grouped|fund, got %r" % tier)
    headers = _CCH_HEADERS[tier]
    numeric = _CCH_NUMERIC[tier]
    textcols = _CCH_TEXT[tier]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)

    total = 0.0
    by_fund = {}
    for ri, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            val = row.get(_CCH_KEYS[h], "")
            if h in ("Account Balance", "Prior Year Balance") and val not in ("", None):
                val = float(val)
            cell = ws.cell(row=ri, column=c, value=val)
            if c in numeric:
                cell.number_format = _CCH_NUMFMT
            elif c in textcols:
                cell.number_format = "@"
        bal = row.get("balance", 0) or 0
        total += float(bal)
        if tier == "fund":
            f = str(row.get("fund_index", "") or "")
            by_fund[f] = by_fund.get(f, 0.0) + float(bal)

    # HARD RULE: CCH silently rejects a workbook with freeze panes.
    ws.freeze_panes = None
    wb.save(out_path)
    return {"tier": tier, "rows": len(rows), "total": round(total, 2),
            "by_fund": {k: round(v, 2) for k, v in by_fund.items()}}


def write_fund_list(out_path, funds, sheet_name="Funds"):
    """Write a standalone fund-definition import (optional deliverable).

    funds: list of dicts with keys fund_index, fund_name, fund_type.
    Sterile, no freeze panes — same constraints as the TB import.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ["Fund Index", "Fund Name", "Fund Type"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    for ri, f in enumerate(funds, 2):
        ws.cell(row=ri, column=1, value=str(f.get("fund_index", "") or "")).number_format = "@"
        ws.cell(row=ri, column=2, value=f.get("fund_name", ""))
        ws.cell(row=ri, column=3, value=f.get("fund_type", ""))
    ws.freeze_panes = None
    wb.save(out_path)
    return {"funds": len(funds)}


# --------------------------------------------------------------------------
# CLI: quick inspection — `python tb_io.py inspect <tb> [<legend>]`
# --------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 3 and sys.argv[1] == "inspect":
        tb = load_tb(sys.argv[2])
        codes = [r["code"] for r in tb]
        print("TB rows:", len(tb))
        print("Code format:", detect_format(codes))
        uniq = sorted({norm_code(c) for c in codes if norm_code(c)})
        print("Unique codes (%d):" % len(uniq), ", ".join(uniq))
        if len(sys.argv) >= 4:
            legend = load_legend(sys.argv[3])
            print("\nLegend parsed from", sys.argv[3], "(%d entries):" % len(legend))
            for k in sorted(legend):
                print("  %-6s -> %s" % (k, legend[k] or "(no label)"))
        fp = scan_fund_patterns([r["account"] for r in tb])
        print("\nFund inference: pattern=%s coverage=%s funds=%s"
              % (fp["pattern"], fp["coverage"], ", ".join(fp["funds"]) or "(none)"))
    else:
        print("usage: python tb_io.py inspect <tb_path> [<legend_path>]")
